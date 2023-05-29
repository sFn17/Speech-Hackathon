import speech_recognition as sr
import openpyxl


def transcribe_speech():
    recognizer = sr.Recognizer()
    with sr.Microphone() as source:
        print("Începeți să vorbiți...")
        audio = recognizer.listen(source)
    try:
        text = recognizer.recognize_google(audio, language='ro-RO')
        return text
    except sr.UnknownValueError:
        print("Nu am putut recunoaște discursul.")
    except sr.RequestError:
        print("Eroare la solicitarea serviciului de recunoaștere a discursului.")


def get_answer(question, file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    for row in sheet.iter_rows(values_only=True):
        if row[0] == question:
            return row[1]
    return "Nu am găsit un răspuns pentru această întrebare."


excel_file_path = 'C:\\Users\\rares\\Desktop\\Speech\\rasp.xlsx'


wb = openpyxl.load_workbook(excel_file_path)
sheet = wb.active
questions = [row[0] for row in sheet.iter_rows(values_only=True)]


while True:
    question = transcribe_speech()
    if question:
        if question.lower() == 'stop':
            break
        if question in questions:
            answer = get_answer(question, excel_file_path)
            print("Întrebare:", question)
            print("Răspuns:", answer)
        else:
            print("Întrebare invalidă.")
